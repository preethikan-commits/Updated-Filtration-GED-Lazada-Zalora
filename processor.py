import pandas as pd
import numpy as np
import re
from datetime import datetime
import io


# ─────────────────────────────────────────────
# File Loading
# ─────────────────────────────────────────────

def load_file(uploaded_file, sheet_name=0, header=0, dtype=None):
    """Load an uploaded Streamlit file (xlsx/xls/csv) into a DataFrame."""
    name = uploaded_file.name.lower()
    uploaded_file.seek(0)
    if name.endswith(".csv"):
        return pd.read_csv(uploaded_file, dtype=dtype)
    else:
        return pd.read_excel(uploaded_file, sheet_name=sheet_name, header=header, dtype=dtype)


def _detect_ecom_header_row(raw_df, max_scan=10):
    """
    Scan the first max_scan rows of a raw (header=None) DataFrame to find
    the row containing the Ecom Tracker header (identified by STYLE# or common
    marker columns). Returns the 0-based row index, or 3 as fallback.
    """
    markers = {"style#", "style #", "style no", "sku", "article"}
    for i in range(min(max_scan, len(raw_df))):
        row_vals = [str(v).strip().lower() for v in raw_df.iloc[i] if pd.notna(v)]
        if any(m in row_vals for m in markers):
            return i
    return 3  # documented default: headers on row 4 (0-indexed: 3)


def _load_sheet_raw(uploaded_file, sheet_name):
    """Load a sheet with header=None to allow manual header detection."""
    uploaded_file.seek(0)
    return pd.read_excel(uploaded_file, sheet_name=sheet_name, header=None)


def load_ecom_tracker(uploaded_file, region):
    """
    Load Ecom Tracker, auto-detecting the true header row by scanning for STYLE#.
    Falls back to row 4 (index 3) per spec if not found.
    Tries the region-specific sheet first, then falls back to the first sheet.
    """
    uploaded_file.seek(0)
    name = uploaded_file.name.lower()

    if name.endswith(".csv"):
        raw = pd.read_csv(uploaded_file, header=None)
        header_row = _detect_ecom_header_row(raw)
        uploaded_file.seek(0)
        return pd.read_csv(uploaded_file, header=header_row)

    sheet_to_use = None
    try:
        raw = _load_sheet_raw(uploaded_file, region)
        sheet_to_use = region
    except Exception:
        try:
            uploaded_file.seek(0)
            raw = _load_sheet_raw(uploaded_file, 0)
            sheet_to_use = 0
        except Exception:
            uploaded_file.seek(0)
            raw = pd.read_excel(uploaded_file, header=None)
            sheet_to_use = 0

    header_row = _detect_ecom_header_row(raw)
    uploaded_file.seek(0)
    df = pd.read_excel(uploaded_file, sheet_name=sheet_to_use, header=header_row)
    return df


def load_am_exclusion(uploaded_file, region):
    """Load AM Exclusion sheet based on region."""
    sheet_map = {"SG": "SG VC Exclusions", "MY": "MY VC Exclusions"}
    sheet_name = sheet_map.get(region, 0)
    uploaded_file.seek(0)
    name = uploaded_file.name.lower()
    if name.endswith(".csv"):
        return pd.read_csv(uploaded_file)
    try:
        df = pd.read_excel(uploaded_file, sheet_name=sheet_name)
    except Exception:
        uploaded_file.seek(0)
        df = pd.read_excel(uploaded_file, sheet_name=0)
    return df


# ─────────────────────────────────────────────
# Column Utilities
# ─────────────────────────────────────────────

def col_letter_to_index(letter):
    """Convert Excel column letter(s) to 0-based index. e.g. 'A'->0, 'AX'->49"""
    letter = letter.upper()
    idx = 0
    for ch in letter:
        idx = idx * 26 + (ord(ch) - ord('A') + 1)
    return idx - 1


def find_col(df, candidates):
    """Return the first column name from df that matches any candidate (case-insensitive)."""
    cols_lower = {c.strip().lower(): c for c in df.columns}
    for cand in candidates:
        key = cand.strip().lower()
        if key in cols_lower:
            return cols_lower[key]
    return None


def normalize_str(s):
    if pd.isna(s):
        return ""
    return str(s).strip()


# ─────────────────────────────────────────────
# Price Tier Column Parsing
# ─────────────────────────────────────────────

def _find_tier_col_by_name(tier_df_cols, candidates):
    """Case-insensitive name search within a list of column names."""
    lower_map = {c.strip().lower(): c for c in tier_df_cols}
    for cand in candidates:
        key = cand.strip().lower()
        if key in lower_map:
            return lower_map[key]
    return None


def parse_price_tier_ref(ref, ecom_df):
    """
    Resolve price tier columns from a reference like 'AX-BA'.

    Two-pass strategy:
      Pass 1 — name-based: search for RRP/SRP/DISC%/Exclusion header names.
      Pass 2 — positional fallback using absolute or relative indices.

    Fixed order within the range:
        START   = RRP
        START+1 = SRP
        START+2 = DISC %
        END     = Exclusion (always last column in range)
    """
    ref = ref.strip().upper()
    if "-" not in ref:
        raise ValueError(f"Invalid Price Tier Reference: '{ref}'. Expected format like 'AX-BA'.")

    parts = ref.split("-")
    start_letter = parts[0].strip()
    end_letter   = parts[1].strip()

    start_abs = col_letter_to_index(start_letter)
    end_abs   = col_letter_to_index(end_letter)

    cols   = list(ecom_df.columns)
    n_cols = len(cols)

    # Pass 1: name-based
    rrp_col       = _find_tier_col_by_name(cols, ["RRP", "Retail Price", "Retail Selling Price"])
    srp_col       = _find_tier_col_by_name(cols, ["SRP", "Selling Price", "Sale Price"])
    disc_col      = _find_tier_col_by_name(cols, ["DISC %", "DISC%", "Discount %", "Discount%", "Disc %"])
    exclusion_col = _find_tier_col_by_name(cols, ["Exclusion", "Exclusion Remarks", "Excl", "VC Exclusion", "Voucher Exclusion"])

    # Pass 2: positional fallback
    span = end_abs - start_abs + 1
    if not all([rrp_col, srp_col, disc_col, exclusion_col]):
        # Strategy a: absolute indices fit
        if start_abs < n_cols and end_abs < n_cols:
            tier_cols_pos = cols[start_abs: end_abs + 1]
            if len(tier_cols_pos) >= 4:
                if not rrp_col:       rrp_col       = tier_cols_pos[0]
                if not srp_col:       srp_col       = tier_cols_pos[1]
                if not disc_col:      disc_col      = tier_cols_pos[2]
                if not exclusion_col: exclusion_col = tier_cols_pos[-1]

        # Strategy b: last N columns
        if not all([rrp_col, srp_col, disc_col, exclusion_col]):
            tail = cols[max(0, n_cols - span):]
            if len(tail) >= 4:
                if not rrp_col:       rrp_col       = tail[0]
                if not srp_col:       srp_col       = tail[1]
                if not disc_col:      disc_col      = tail[2]
                if not exclusion_col: exclusion_col = tail[-1]

        # Strategy c: last 4 cols
        if not all([rrp_col, srp_col, disc_col, exclusion_col]):
            if n_cols >= 4:
                offset = max(0, n_cols - 4)
                if not rrp_col:       rrp_col       = cols[offset]
                if not srp_col:       srp_col       = cols[offset + 1]
                if not disc_col:      disc_col      = cols[offset + 2]
                if not exclusion_col: exclusion_col = cols[-1]
            else:
                raise ValueError(
                    f"Ecom Tracker has only {n_cols} column(s) — cannot extract 4 price tier columns. "
                    f"Please check the correct sheet/file is uploaded and headers start at row 4."
                )

    try:
        s = cols.index(rrp_col)
        e = cols.index(exclusion_col)
        tier_cols = cols[s: e + 1]
    except ValueError:
        tier_cols = [rrp_col, srp_col, disc_col, exclusion_col]

    return {
        "tier_cols":      tier_cols,
        "rrp_col":        rrp_col,
        "srp_col":        srp_col,
        "disc_col":       disc_col,
        "exclusion_col":  exclusion_col,
        "start_idx":      start_abs,
        "end_idx":        end_abs,
        "start_letter":   start_letter,
        "end_letter":     end_letter,
    }


# ─────────────────────────────────────────────
# Date Processing
# ─────────────────────────────────────────────
#
# The Ecom Tracker stores Launch Dates in DD-MM-YYYY format.
# When pandas reads an Excel date column it converts cells to Python datetime
# objects, so we handle that case first before falling back to string parsing.

INVALID_DATE_TEXTS = {"past season", "tbc", "00-jan-1900", ""}

def parse_launch_date(val):
    """
    Return a datetime.date or None.

    Priority order:
      1. Already a date/datetime object (pandas Excel read) → use directly.
      2. String in DD-MM-YYYY format (primary Ecom Tracker format) → parse first.
      3. Other common date strings → try remaining formats.
      4. Reject: blank, text-only, 00-Jan-1900, 'Past Season', 'TBC', etc.
    """
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None

    # ── Path 1: already a Python date / datetime (pandas Excel conversion) ──
    if isinstance(val, datetime):
        # Guard against Excel's epoch ghost date 1900-01-00
        if val.year == 1900 and val.month == 1 and val.day <= 1:
            return None
        return val.date()
    if hasattr(val, 'date') and callable(val.date):
        # covers pandas Timestamp and datetime subclasses
        try:
            d = val.date()
            if d.year == 1900 and d.month == 1 and d.day <= 1:
                return None
            return d
        except Exception:
            pass

    # ── Path 2: string parsing ──
    try:
        is_na = pd.isna(val)
    except Exception:
        is_na = False
    if is_na:
        return None

    s = str(val).strip()
    if not s or s.lower() in INVALID_DATE_TEXTS:
        return None
    # Reject pure-text values (no digit = not a date)
    if not any(ch.isdigit() for ch in s):
        return None
    # Reject Excel ghost date variants
    if "1900" in s and ("jan" in s.lower() or s.startswith("00")):
        return None

    # Try DD-MM-YYYY first (primary Ecom Tracker format), then others
    for fmt in (
        "%d-%m-%Y",   # ← PRIMARY: DD-MM-YYYY (Ecom Tracker)
        "%d/%m/%Y",   # DD/MM/YYYY
        "%Y-%m-%d",   # ISO (pandas default str representation)
        "%d %b %Y",   # 13 May 2025
        "%d-%b-%Y",   # 13-May-2025
        "%d/%b/%Y",   # 13/May/2025
        "%m/%d/%Y",   # US format fallback
    ):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass

    # Last resort: let pandas infer (explicit formats above already cover DD-MM-YYYY)
    try:
        return pd.to_datetime(s, dayfirst=False).date()
    except Exception:
        pass

    return None


def format_launch_date(val):
    """
    Return the launch date as a DD-MM-YYYY string, or empty string if invalid.
    Preserves the Ecom Tracker's native DD-MM-YYYY format throughout.
    """
    d = parse_launch_date(val)
    if d is None:
        return ""
    return d.strftime("%d-%m-%Y")


# ─────────────────────────────────────────────
# Voucher Percentage Parsing
# ─────────────────────────────────────────────

def extract_voucher_pct(name):
    """
    Extract the leading numeric percentage from a voucher name.
    Supports: '10% VC' -> 10.0, '30 % NMS' -> 30.0, '40 percent' -> 40.0
    Returns None if no percentage found.
    """
    m = re.search(r"(\d+(?:\.\d+)?)\s*(?:%|percent)", name.strip(), re.IGNORECASE)
    if m:
        return float(m.group(1))
    return None


# ─────────────────────────────────────────────
# AM Exclusion Logic  (fully rewritten)
# ─────────────────────────────────────────────
#
# Rule 1 — Global exclusion keywords → exclude from ALL vouchers
# Rule 2 — "Exclude from X%" → exclude from vouchers with exactly X%
# Rule 3 — "Exclude from X% and above" → exclude from vouchers with % >= X
#
# Multiple rules can apply to one ALU; all are applied (most restrictive wins).

# Global exclusion keyword fragments (case-insensitive substring match)
_GLOBAL_EXCL_KEYWORDS = [
    "exclude from all voucher",
    "exclude from platform voucher",
    "exclude from vc",
    "exclude from voucher",
    "voucher exclusion",
]

# Bare "Exclude" or "Exclude from VC" with no percentage
_BARE_EXCLUDE_RE = re.compile(
    r"^\s*exclude\s*$"
    r"|exclude\s+from\s+vc\s*$"
    r"|exclude\s+from\s+voucher\s*$",
    re.IGNORECASE,
)

# "Exclude from X% and above" / "Exclude from X% above"
_EXCL_GTE_RE = re.compile(
    r"exclude.*?(\d+(?:\.\d+)?)\s*(?:%|percent).*?\b(and\s+)?above\b",
    re.IGNORECASE,
)

# "Exclude from X%" (specific, no "above") — catches "Exclude from 10% VC",
# "Exclude from 15%", "Exclude from 20% Voucher" etc.
_EXCL_EXACT_RE = re.compile(
    r"exclude.*?(\d+(?:\.\d+)?)\s*(?:%|percent)",
    re.IGNORECASE,
)


def parse_am_exclusion_rules(excl_text):
    """
    Parse an AM Exclusion cell value and return a list of rule dicts.
    Each rule has the form: {"mode": "all"|"exact"|"gte", "threshold": float|None}

    Multiple rules may apply to a single ALU — caller applies them all.
    """
    if pd.isna(excl_text) or not str(excl_text).strip():
        return []

    text = str(excl_text).strip()
    rules = []

    # ── Rule 1: global exclusion (keyword match) ──
    text_lower = text.lower()
    is_global = any(kw in text_lower for kw in _GLOBAL_EXCL_KEYWORDS)
    if not is_global and _BARE_EXCLUDE_RE.search(text):
        is_global = True
    if is_global:
        return [{"mode": "all", "threshold": None}]

    # ── Rule 3: "X% and above" ── (must check before exact to avoid false match)
    for m in _EXCL_GTE_RE.finditer(text):
        rules.append({"mode": "gte", "threshold": float(m.group(1))})

    # ── Rule 2: exact percentage (lines that do NOT also contain "above") ──
    # Strip already-matched "and above" segments before scanning for exact
    text_no_above = _EXCL_GTE_RE.sub("", text)
    for m in _EXCL_EXACT_RE.finditer(text_no_above):
        pct = float(m.group(1))
        # Avoid duplicating a threshold already captured by Rule 3
        already_gte = any(r["threshold"] == pct and r["mode"] == "gte" for r in rules)
        if not already_gte:
            rules.append({"mode": "exact", "threshold": pct})

    return rules


def is_am_excluded_for_voucher(rules, voucher_pct):
    """
    Given a list of parsed AM exclusion rules and the voucher's numeric %,
    return True if the ALU should be excluded from that voucher.

    If voucher_pct is None (no % in voucher name), only global rules apply.
    """
    for rule in rules:
        mode      = rule["mode"]
        threshold = rule["threshold"]

        if mode == "all":
            return True
        if mode == "exact" and voucher_pct is not None and voucher_pct == threshold:
            return True
        if mode == "gte" and voucher_pct is not None and voucher_pct >= threshold:
            return True

    return False


def am_exclusion_label(rules):
    """Return a human-readable label for the AM Exclude column from a rule list."""
    if not rules:
        return ""
    if any(r["mode"] == "all" for r in rules):
        return "Exclude from all Vouchers"
    parts = []
    for r in rules:
        if r["mode"] == "exact":
            t = int(r["threshold"]) if r["threshold"] == int(r["threshold"]) else r["threshold"]
            parts.append(f"Exclude from {t}%")
        elif r["mode"] == "gte":
            t = int(r["threshold"]) if r["threshold"] == int(r["threshold"]) else r["threshold"]
            parts.append(f"Exclude from {t}% and above")
    return " | ".join(parts) if parts else ""


# ─────────────────────────────────────────────
# Exclusion Remarks Extraction (for dropdown)
# ─────────────────────────────────────────────

def get_exclusion_remarks(ecom_df, exclusion_col):
    """
    Extract unique, non-blank values from the Exclusion column of the Ecom Tracker.
    Used to populate the inclusion keyword dropdown in the UI.
    Returns a sorted list of strings.
    """
    if not exclusion_col or exclusion_col not in ecom_df.columns:
        return []
    vals = (
        ecom_df[exclusion_col]
        .dropna()
        .astype(str)
        .str.strip()
        .replace("", pd.NA)
        .dropna()
        .unique()
        .tolist()
    )
    return sorted(vals)


# ─────────────────────────────────────────────
# Core Processing Function
# ─────────────────────────────────────────────

def process_voucher_eligibility(
    marketplace,
    region,
    sc_report_file,
    ecom_file,
    content_file,
    am_excl_file,
    price_tier_ref,
    cutoff_date,
    voucher_configs,
):
    """
    Main processing function. Returns (output_df, warnings_list).

    Each voucher in voucher_configs is evaluated completely independently:
      1. Launch Date filter
      2. Ecom Status filter
      3. RRP filter
      4. SRP filter
      5. Inclusion Keyword filter (exact match against Exclusion column values)
      6. AM Exclusion filter (per-voucher, using parsed rules)

    Eligible SKUs are marked "Yes-Eligible"; ineligible SKUs are left blank.
    """
    warnings = []

    # ── 1. Load SC Report ──
    sc_df = load_file(sc_report_file, dtype=str)
    sc_df.columns = sc_df.columns.str.strip()

    # ── 2. Load Content File ──
    content_df = load_file(content_file, dtype=str)
    content_df.columns = content_df.columns.str.strip()

    # ── 3. Load Ecom Tracker ──
    ecom_df = load_ecom_tracker(ecom_file, region)
    ecom_df.columns = [str(c).strip() for c in ecom_df.columns]

    # ── 4. Load AM Exclusion (optional) ──
    am_df = None
    if am_excl_file is not None:
        am_df = load_am_exclusion(am_excl_file, region)
        am_df.columns = am_df.columns.str.strip()

    # ── 5. Parse Price Tier Reference ──
    tier_info     = parse_price_tier_ref(price_tier_ref, ecom_df)
    rrp_col       = tier_info["rrp_col"]
    srp_col       = tier_info["srp_col"]
    disc_col      = tier_info["disc_col"]
    exclusion_col = tier_info["exclusion_col"]

    # Info message showing resolved column mapping
    cols_list = list(ecom_df.columns)
    def _col_label(col_name):
        try:
            idx = cols_list.index(col_name)
            n, letters = idx + 1, ""
            while n:
                n, r = divmod(n - 1, 26)
                letters = chr(65 + r) + letters
            return f"{letters} ({col_name})"
        except ValueError:
            return col_name

    warnings.append(
        f"ℹ️ Price Tier column mapping — "
        f"RRP: {_col_label(rrp_col)} | "
        f"SRP: {_col_label(srp_col)} | "
        f"DISC %: {_col_label(disc_col)} | "
        f"Exclusion: {_col_label(exclusion_col)}"
    )

    # ── 6. Identify key columns in Content File ──
    ean_col         = find_col(content_df, ["EAN", "EAN Code", "EAN Number"])
    alu_col_content = find_col(content_df, ["Color No", "Colour No", "Color No.", "Color Number", "ALU"])
    if not ean_col:
        raise ValueError("Content File: Cannot find 'EAN' column.")
    if not alu_col_content:
        raise ValueError("Content File: Cannot find 'Color No' / ALU column.")

    content_map = (
        content_df[[ean_col, alu_col_content]]
        .dropna(subset=[ean_col])
        .drop_duplicates(subset=[ean_col])
        .set_index(ean_col)[alu_col_content]
        .to_dict()
    )

    # ── 7. Identify SellerSKU column in SC Report ──
    seller_sku_col = find_col(sc_df, ["SellerSKU", "Seller SKU", "SKU", "seller_sku"])
    if not seller_sku_col:
        raise ValueError("SC Report: Cannot find 'SellerSKU' column.")

    # ── 8. Map ALU onto SC Report ──
    sc_df["ALU"] = sc_df[seller_sku_col].map(content_map).fillna("")

    # ── 9. Identify key columns in Ecom Tracker ──
    style_col = find_col(ecom_df, ["STYLE#", "Style#", "STYLE #", "Style No", "StyleNo", "ALU"])
    if not style_col:
        raise ValueError("Ecom Tracker: Cannot find 'STYLE#' column.")

    launch_date_col = find_col(
        ecom_df,
        ["Launch Dates", "Launch Date", "LaunchDate", "Launch date",
         "LAUNCH DATE", "LAUNCH DATES", "Ecom Launch Date"],
    )
    if not launch_date_col:
        warnings.append("Ecom Tracker: 'Launch Dates' column not found — Launch Date will be blank.")

    if marketplace == "Lazada":
        ecom_status_col = find_col(ecom_df, ["Lazada", "LAZADA", "Lazada Status", "Lazada Ecom Status"])
    else:
        ecom_status_col = find_col(ecom_df, ["Zalora", "ZALORA", "Zalora Status", "Zalora Ecom Status"])

    if not ecom_status_col:
        warnings.append(
            f"Ecom Tracker: '{marketplace}' column not found — Ecom Status will be blank. "
            f"Expected a column named '{marketplace}'."
        )

    # ── 10. Build ALU → Ecom Tracker row lookup ──
    ecom_df[style_col] = ecom_df[style_col].astype(str).str.strip()
    ecom_lookup = ecom_df.drop_duplicates(subset=[style_col]).set_index(style_col)

    def get_ecom_val(alu, col):
        if not col or not alu or alu not in ecom_lookup.index:
            return ""
        val = ecom_lookup.at[alu, col]
        return "" if pd.isna(val) else val

    # ── 11. AM Exclusion mapping: ALU → list of rule dicts ──
    am_excl_map = {}   # ALU → [{"mode":..., "threshold":...}, ...]
    if am_df is not None:
        article_col   = find_col(am_df, ["Article", "ALU", "SKU", "Style No"])
        excl_type_col = find_col(am_df, ["Exclusion Type", "ExclusionType", "Exclusion", "Type"])
        if article_col and excl_type_col:
            for _, row in am_df.iterrows():
                alu_key   = normalize_str(row[article_col])
                excl_text = normalize_str(row[excl_type_col])
                if alu_key:
                    rules = parse_am_exclusion_rules(excl_text)
                    if alu_key in am_excl_map:
                        am_excl_map[alu_key].extend(rules)   # merge multiple rows
                    else:
                        am_excl_map[alu_key] = rules

    # ── 12. Build output DataFrame ──
    out  = sc_df.copy()
    alus = out["ALU"].astype(str).str.strip()

    # Launch Date
    out["Launch Date"] = (
        alus.map(lambda a: get_ecom_val(a, launch_date_col)).map(format_launch_date)
        if launch_date_col else ""
    )

    # Ecom Status
    out["Ecom Status"] = (
        alus.map(lambda a: normalize_str(get_ecom_val(a, ecom_status_col)))
        if ecom_status_col else ""
    )

    # Price fields
    out["RRP"]    = alus.map(lambda a: get_ecom_val(a, rrp_col))    if rrp_col    else ""
    out["SRP"]    = alus.map(lambda a: get_ecom_val(a, srp_col))    if srp_col    else ""
    out["DISC %"] = alus.map(lambda a: get_ecom_val(a, disc_col))   if disc_col   else ""

    # Ecom Tracker Exclusion / Inclusion Remarks
    out["Exclusion"] = alus.map(lambda a: normalize_str(get_ecom_val(a, exclusion_col)))

    # AM Exclude — human-readable label
    out["AM Exclude"] = alus.map(
        lambda a: am_exclusion_label(am_excl_map.get(a, []))
    )

    # ── 13. Per-voucher independent eligibility evaluation ──
    def _to_float(val):
        try:
            return float(str(val).replace(",", "").strip())
        except Exception:
            return None

    def evaluate_sku_for_voucher(row, keywords, vc_pct):
        """
        Evaluate one SKU row against one voucher's full criteria independently.
        Returns "Yes-Eligible" if all filters pass, "" otherwise.

        Filter order:
          1. Launch Date   — valid, not future
          2. Ecom Status   — must be YES
          3. RRP           — must be > 16
          4. SRP           — must be 0 or > 16
          5. Inclusion     — Exclusion column must exactly match a voucher keyword
          6. AM Exclusion  — ALU must not be excluded for this voucher %
        """
        alu = str(row.get("ALU", "")).strip()

        # Filter 1: Launch Date
        raw_ld = str(row.get("Launch Date", "")).strip()
        if not raw_ld:
            return ""
        ld = parse_launch_date(raw_ld)
        if ld is None:
            return ""
        if cutoff_date and ld > cutoff_date:
            return ""

        # Filter 2: Ecom Status
        if str(row.get("Ecom Status", "")).strip().upper() != "YES":
            return ""

        # Filter 3: RRP
        rrp = _to_float(row.get("RRP", ""))
        if rrp is None or rrp <= 16:
            return ""

        # Filter 4: SRP (0 is allowed; 1–16 inclusive is excluded)
        srp = _to_float(row.get("SRP", ""))
        if srp is not None and srp != 0 and srp <= 16:
            return ""

        # Filter 5: Inclusion Keyword (exact match, case-sensitive)
        ecom_excl_val = normalize_str(row.get("Exclusion", ""))
        if keywords:
            if not any(ecom_excl_val == kw for kw in keywords):
                return ""

        # Filter 6: AM Exclusion
        if alu:
            rules = am_excl_map.get(alu, [])
            if is_am_excluded_for_voucher(rules, vc_pct):
                return ""

        return "Yes-Eligible"

    # ── 14. Build one result column per voucher ──
    for vc in voucher_configs:
        vc_name  = vc.get("name", "").strip()
        keywords = [kw.strip() for kw in vc.get("keywords", []) if str(kw).strip()]
        vc_pct   = extract_voucher_pct(vc_name)

        out[vc_name] = out.apply(
            lambda row, _kw=keywords, _pct=vc_pct:
                evaluate_sku_for_voucher(row, _kw, _pct),
            axis=1,
        )

    return out, warnings


# ─────────────────────────────────────────────
# Excel Output Generation
# ─────────────────────────────────────────────

WORKING_COLS = ["ALU", "Launch Date", "Ecom Status", "RRP", "SRP", "DISC %", "Exclusion", "AM Exclude"]


def generate_excel_output(df, marketplace, voucher_configs):
    """Return bytes of a styled Excel workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = f"{marketplace} VC Eligibility"

    voucher_names = [vc["name"] for vc in voucher_configs if vc.get("name", "").strip()]

    injected = set(WORKING_COLS + voucher_names)
    sc_cols  = [c for c in df.columns if c not in injected]
    all_cols = sc_cols + WORKING_COLS + voucher_names

    # Styles
    header_font         = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    sc_header_fill      = PatternFill("solid", start_color="2D4A6B")   # Dark blue
    working_header_fill = PatternFill("solid", start_color="1F6B3E")   # Dark green
    vc_header_fill      = PatternFill("solid", start_color="7B2D8B")   # Purple
    eligible_fill       = PatternFill("solid", start_color="C6EFCE")   # Light green
    eligible_font       = Font(name="Arial", bold=True, color="276221", size=10)
    alt_fill            = PatternFill("solid", start_color="F2F2F2")   # Light grey alternate
    cell_font           = Font(name="Arial", size=10)
    center_align        = Alignment(horizontal="center", vertical="center", wrap_text=False)
    left_align          = Alignment(horizontal="left", vertical="center")
    thin                = Side(border_style="thin", color="D0D0D0")
    border              = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Headers
    for ci, col in enumerate(all_cols, start=1):
        cell           = ws.cell(row=1, column=ci, value=col)
        cell.font      = header_font
        cell.alignment = center_align
        if col in sc_cols:
            cell.fill = sc_header_fill
        elif col in WORKING_COLS:
            cell.fill = working_header_fill
        else:
            cell.fill = vc_header_fill

    # Data rows
    for ri, (_, row) in enumerate(df[all_cols].iterrows(), start=2):
        is_alt = (ri % 2 == 0)
        for ci, col in enumerate(all_cols, start=1):
            val = row[col]
            if pd.isna(val):
                val = ""
            cell        = ws.cell(row=ri, column=ci, value=val)
            cell.border = border

            if col in voucher_names and str(val).strip().lower() == "yes-eligible":
                cell.fill      = eligible_fill
                cell.font      = eligible_font
                cell.alignment = center_align
            else:
                cell.font      = cell_font
                cell.alignment = left_align if col in sc_cols else center_align
                if is_alt and col not in voucher_names:
                    cell.fill = alt_fill

    # Auto-width (capped at 45)
    for ci, col in enumerate(all_cols, start=1):
        max_len = len(str(col))
        for ri in range(2, min(ws.max_row + 1, 200)):
            v = ws.cell(row=ri, column=ci).value
            if v:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 45)

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
